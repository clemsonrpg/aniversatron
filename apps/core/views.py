from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Prefetch
from apps.servicos.models import Servico
from apps.pessoas.models import Pessoa
from django.shortcuts import render
import json
from django.db.models.functions import TruncMonth
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from apps.pessoas.models import Propriedade

def index(request):
    hoje = timezone.localdate()

    # OTIMIZAÇÃO: Buscamos os aniversariantes já trazendo o último serviço de cada um
    # Isso evita o problema de "N+1 queries" (consultas dentro do loop)
    aniversariantes_queryset = Pessoa.objects.filter(
        data_nascimento__month=hoje.month,
        data_nascimento__day=hoje.day
    ).prefetch_related(
        Prefetch(
            'servicos',
            queryset=Servico.objects.order_by('-data_servico', '-id'),
            to_attr='ultimos_servicos'
        )
    )

    lista_aniversariantes = []
    for p in aniversariantes_queryset:
        # Cálculo de idade simplificado
        idade = hoje.year - p.data_nascimento.year - (
            (hoje.month, hoje.day) < (p.data_nascimento.month, p.data_nascimento.day)
        )
        
        # Pega o primeiro serviço da lista que prebuscamos (to_attr)
        ultimo_servico = p.ultimos_servicos[0].nome_servico if p.ultimos_servicos else "Nenhum serviço"

        lista_aniversariantes.append({
            "nome": p.nome,
            "email": p.email or "Não informado",
            "telefone": p.telefone or "Não informado",
            "idade": idade,
            "servico": ultimo_servico
        })

    context = {
        'total_pessoas': Pessoa.objects.count(),
        'total_servicos': Servico.objects.count(),
        'lista_aniversariantes': lista_aniversariantes,
    }
    return render(request, 'index.html', context)


def relatorios(request):
    template_name = 'core/relatorios.html'
    hoje = timezone.localdate()
    localidade_filtro = request.GET.get('localidade')
    
    # 1. Base de dados
    servicos_qs = Servico.objects.all()
    
    # 2. Define o Título Dinâmico
    if localidade_filtro:
        servicos_qs = servicos_qs.filter(propriedade__localidade=localidade_filtro)
        # Busca o nome da localidade nos choices
        dict_choices = dict(Propriedade._meta.get_field('localidade').choices)
        localidade_nome_exibicao = dict_choices.get(localidade_filtro)
        titulo_painel = f"Evolução Mensal - {localidade_nome_exibicao}"
        subtitulo_lista = f"Serviços Pendentes - {localidade_nome_exibicao}"
    else:
        localidade_nome_exibicao = None
        titulo_painel = "Evolução Mensal - Geral"
        subtitulo_lista = "Todos os Serviços Pendentes"

    # 3. Separação por Status (Se localidade_filtro for vazio, traz de todas)
    servicos_pendentes = servicos_qs.filter(status="Pendente").select_related('propriedade__pessoa', 'propriedade')
    servicos_realizados = servicos_qs.exclude(status="Pendente")

    # 4. Totais e Ranking
    total_mes = servicos_realizados.filter(data_servico__month=hoje.month, data_servico__year=hoje.year).count()
    total_pendentes = servicos_pendentes.count()
    total_geral_realizado = servicos_realizados.count()

    ranking_servicos = (
        servicos_realizados
        .values('nome_servico')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )

    # 5. Gráfico de Evolução (6 meses)
    ultimos_6_meses = (
        servicos_realizados
        .annotate(mes=TruncMonth('data_servico'))
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('-mes')[:6]
    )
    
    dados_grafico = {
        "labels": [item['mes'].strftime("%m/%Y") for item in reversed(ultimos_6_meses)],
        "valores": [item['total'] for item in reversed(ultimos_6_meses)]
    }

    # 6. Gráfico Donut (Sempre mostra a visão geral por região)
    dict_choices_all = dict(Propriedade._meta.get_field('localidade').choices)
    dados_brutos = Servico.objects.values('propriedade__localidade').annotate(total=Count('id'))
    distribuicao_regiao = [
        {'label_localidade': dict_choices_all.get(item['propriedade__localidade'], item['propriedade__localidade']), 
         'total': item['total']} for item in dados_brutos
    ]

    
    context = {
        "total_mes": total_mes,
        "total_pendentes": total_pendentes,
        "total_servicos_filtrados": total_geral_realizado,
        "ranking_servicos": ranking_servicos,
        "grafico_dados": json.dumps(dados_grafico),
        "distribuicao_regiao": distribuicao_regiao,

            "distribuicao_json": json.dumps([
        {
            "label": item["label_localidade"],
            "total": item["total"]
        }
        for item in distribuicao_regiao
    ]),
        "lista_localidades_choices": Propriedade._meta.get_field('localidade').choices,
        "localidade_nome_exibicao": localidade_nome_exibicao,
        "titulo_painel": titulo_painel,
        "subtitulo_lista": subtitulo_lista,
        "servicos_pendentes": servicos_pendentes,
        "servicos" : servicos_qs,
        "pessoas": Pessoa.objects.all()
    }
    return render(request, template_name, context)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count
from django.db.models.functions import TruncMonth
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from apps.servicos.models import Servico


def relatorio_excel(request):
    hoje = timezone.localdate()

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="relatorio_servicos.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Serviços"

    # =========================
    # DADOS
    # =========================

    total_geral = Servico.objects.count()

    concluidos = Servico.objects.exclude(status="Pendente")
    pendentes = Servico.objects.filter(status="Pendente")

    total_concluidos = concluidos.count()
    total_pendentes = pendentes.count()

    total_mes = concluidos.filter(
        data_servico__month=hoje.month,
        data_servico__year=hoje.year
    ).count()

    ranking = (
        concluidos
        .values("nome_servico")
        .annotate(total=Count("id"))
        .order_by("-total")[:5]
    )

    por_mes = (
        concluidos
        .annotate(mes=TruncMonth('data_servico'))
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('mes')
    )

    # =========================
    # CABEÇALHO
    # =========================

    ws["A1"] = "SECRETARIA DE AGRICULTURA"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A3"] = f"Relatório Estatístico - {hoje.strftime('%d/%m/%Y')}"
    ws["A3"].font = Font(bold=True)

    # =========================
    # TOTAIS
    # =========================

    ws["A5"] = "Total Geral:"
    ws["B5"] = total_geral

    ws["A6"] = "Total Concluídos:"
    ws["B6"] = total_concluidos

    ws["A7"] = "Total Pendentes:"
    ws["B7"] = total_pendentes

    ws["A8"] = "Total Concluídos no Mês Atual:"
    ws["B8"] = total_mes

    # =========================
    # RANKING
    # =========================

    ws["A10"] = "Serviços Mais Realizados"
    ws["A10"].font = Font(size=12, bold=True)

    ws["A12"] = "Serviço"
    ws["B12"] = "Total"

    ws["A12"].font = Font(bold=True)
    ws["B12"].font = Font(bold=True)

    linha = 13
    for item in ranking:
        ws[f"A{linha}"] = item["nome_servico"]
        ws[f"B{linha}"] = item["total"]
        linha += 1

    fim_ranking = linha - 1

    # =========================
    # SERVIÇOS POR MÊS
    # =========================

    linha += 2
    ws[f"A{linha}"] = "Serviços Concluídos por Mês"
    ws[f"A{linha}"].font = Font(size=12, bold=True)

    linha += 2
    ws[f"A{linha}"] = "Mês"
    ws[f"B{linha}"] = "Quantidade"

    ws[f"A{linha}"].font = Font(bold=True)
    ws[f"B{linha}"].font = Font(bold=True)

    linha += 1
    inicio_mes = linha

    for item in por_mes:
        ws[f"A{linha}"] = item["mes"].strftime("%m/%Y")
        ws[f"B{linha}"] = item["total"]
        linha += 1

    fim_mes = linha - 1



    # =========================
    # LARGURA DAS COLUNAS
    # =========================

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    wb.save(response)
    return response